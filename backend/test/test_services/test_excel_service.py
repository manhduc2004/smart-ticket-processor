import pytest
import os
from openpyxl import load_workbook

from app.service.excel_service import ExcelService
from app.schemas.ticket import TicketData


class TestExcelService:
    """Test ExcelService."""

    def test_init(self):
        """Test khởi tạo service."""
        service = ExcelService()
        assert service.COLUMNS
        assert len(service.COLUMNS) == 15

    def test_columns_order(self):
        """Test thứ tự columns."""
        service = ExcelService()
        expected = [
            "Airlines",
            "Ticket number",
            "Cust ID",
            "Selling",
            "Commission Selling",
            "Buying",
            "Commission Buying",
            "Đơn vị tiền tệ",
            "Fare",
            "Treo hàng amount",
            "Admin amount",
            "Admin name",
            "Service",
            "Note",
            "Booker name",
        ]
        assert service.COLUMNS == expected

    def test_export_single_ticket(self, temp_export_dir):
        """Test export 1 vé."""
        service = ExcelService()
        
        tickets = [
            TicketData(
                airlines="VNA",
                ticket_number="7383007215202",
                cust_id="TIKET",
                selling=None,
                commission_selling=None,
                buying=9590000,
                commission_buying=0,
                currency="VND",
                fare=9684000,
                admin_amount=None,
                service=None,
                note=None
            )
        ]
        
        filepath = service.export(tickets, output_dir=temp_export_dir)
        
        # Check file exists
        assert os.path.exists(filepath)
        assert filepath.endswith(".xlsx")
        
        # Check content
        wb = load_workbook(filepath)
        ws = wb.active
        
        # Check header
        assert ws.cell(1, 1).value == "Airlines"
        assert ws.cell(1, 2).value == "Ticket number"
        
        # Check data
        assert ws.cell(2, 1).value == "VNA"
        assert ws.cell(2, 2).value == "7383007215202"
        assert ws.cell(2, 3).value == "TIKET"
        assert ws.cell(2, 6).value == 9590000
        assert ws.cell(2, 8).value == "VND"
        assert ws.cell(2, 9).value == 9684000

    def test_export_multiple_tickets(self, temp_export_dir):
        """Test export nhiều vé."""
        service = ExcelService()
        
        tickets = [
            TicketData(
                airlines="VNA",
                ticket_number="1111111111111",
                buying=1000000,
                currency="VND"
            ),
            TicketData(
                airlines="VJA",
                ticket_number="2222222222222",
                buying=2000000,
                currency="VND"
            ),
            TicketData(
                airlines="KAL",
                ticket_number="3333333333333",
                buying=3000000,
                currency="USD"
            ),
        ]
        
        filepath = service.export(tickets, output_dir=temp_export_dir)
        
        wb = load_workbook(filepath)
        ws = wb.active
        
        # Check có 3 vé + 1 header = 4 rows
        assert ws.max_row == 4
        
        # Check data
        assert ws.cell(2, 1).value == "VNA"
        assert ws.cell(3, 1).value == "VJA"
        assert ws.cell(4, 1).value == "KAL"

    def test_export_with_null_values(self, temp_export_dir):
        """Test export với null values."""
        service = ExcelService()
        
        tickets = [
            TicketData(
                airlines=None,
                ticket_number=None,
                cust_id=None,
                buying=None,
                currency=None
            )
        ]
        
        filepath = service.export(tickets, output_dir=temp_export_dir)
        
        wb = load_workbook(filepath)
        ws = wb.active
        
        # Check null values
        assert ws.cell(2, 1).value is None
        assert ws.cell(2, 2).value is None

    def test_export_filename_format(self, temp_export_dir):
        """Test format filename."""
        service = ExcelService()
        
        tickets = [TicketData(airlines="VNA")]
        filepath = service.export(tickets, output_dir=temp_export_dir)
        
        filename = os.path.basename(filepath)
        assert filename.startswith("tickets_")
        assert filename.endswith(".xlsx")
        assert len(filename) == len("tickets_20240101_123456.xlsx")

    def test_get_column_letter(self):
        """Test _get_column_letter helper."""
        service = ExcelService()
        
        assert service._get_column_letter(1) == "A"
        assert service._get_column_letter(2) == "B"
        assert service._get_column_letter(26) == "Z"
        assert service._get_column_letter(27) == "AA"
        assert service._get_column_letter(52) == "AZ"

    def test_export_creates_directory(self):
        """Test tự tạo directory nếu chưa có."""
        service = ExcelService()
        
        output_dir = "/tmp/test_exports_" + str(os.getpid())
        
        try:
            tickets = [TicketData(airlines="VNA")]
            filepath = service.export(tickets, output_dir=output_dir)
            
            assert os.path.exists(output_dir)
            assert os.path.exists(filepath)
        finally:
            # Cleanup
            if os.path.exists(filepath):
                os.remove(filepath)
            if os.path.exists(output_dir):
                os.rmdir(output_dir)