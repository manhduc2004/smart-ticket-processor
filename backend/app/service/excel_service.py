from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from typing import List
from datetime import datetime
import os

from app.schemas.ticket import TicketData


class ExcelService:
    """Service xuất dữ liệu vé ra Excel."""

    COLUMNS = [
        "Airlines",
        "Ticket number",
        "Cust ID",
        "Selling",
        "Commission Selling",
        "Buying",
        "Commission Buying",
        "Đơn vị tiền tệ",
        "Fare",
        "Treo hàng amount",
        "Admin amount",
        "Admin name",
        "Service",
        "Note",
        "Booker name",
    ]

    def export(self, tickets: List[TicketData], output_dir: str = "exports") -> str:
        """Xuất tickets ra file Excel."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Tickets"

        # Header
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)

        for col_idx, col_name in enumerate(self.COLUMNS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Data rows
        for row_idx, ticket in enumerate(tickets, start=2):
            ws.cell(row=row_idx, column=1, value=ticket.airlines)
            ws.cell(row=row_idx, column=2, value=ticket.ticket_number)
            ws.cell(row=row_idx, column=3, value=ticket.cust_id)
            ws.cell(row=row_idx, column=4, value=ticket.selling)
            ws.cell(row=row_idx, column=5, value=ticket.commission_selling)
            ws.cell(row=row_idx, column=6, value=ticket.buying)
            ws.cell(row=row_idx, column=7, value=ticket.commission_buying)
            ws.cell(row=row_idx, column=8, value=ticket.currency)
            ws.cell(row=row_idx, column=9, value=ticket.fare)
            ws.cell(row=row_idx, column=10, value=None)
            ws.cell(row=row_idx, column=11, value=ticket.admin_amount)
            ws.cell(row=row_idx, column=12, value=None)
            ws.cell(row=row_idx, column=13, value=ticket.service)
            ws.cell(row=row_idx, column=14, value=ticket.note)
            ws.cell(row=row_idx, column=15, value=None)

        # Auto-size columns
        for col_idx in range(1, len(self.COLUMNS) + 1):
            ws.column_dimensions[self._get_column_letter(col_idx)].width = 18

        # Save
        os.makedirs(output_dir, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"tickets_{timestamp}.xlsx"
        filepath = os.path.join(output_dir, filename)

        wb.save(filepath)
        return filepath

    @staticmethod
    def _get_column_letter(col_idx: int) -> str:
        """Convert column index → letter."""
        result = ""
        while col_idx > 0:
            col_idx -= 1
            result = chr(col_idx % 26 + ord('A')) + result
            col_idx //= 26
        return result


excel_service = ExcelService()